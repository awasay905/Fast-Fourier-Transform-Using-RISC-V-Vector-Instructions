import openpyxl
from openpyxl.styles import PatternFill

# Function to apply the coloring
def color_spreadsheet(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)

    # Define color shades (HEX)
    base_colors = [
        'FFc9f0d3',  # Column 1
        'FFc9f0d3',  # Column 2
        'FF9ad1c4',  # Column 5
        'FF9ad1c4',  # Column 6
        'FF77b1a9',  # Column 7
        'FF77b1a9',  # Column 8
    ]
    
    # Create fills for columns
    fills = [PatternFill(start_color=color[2:], end_color=color[2:], fill_type="solid") for color in base_colors]

    # Iterate through each spreadsheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Set header color to dark grey
        for col in range(1, 11):  # Columns 1 to 10
            sheet.cell(row=1, column=col).fill = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")

        # Color columns 1 to 8 with specified shades
        for col in range(1, 9):
            for row in range(2, sheet.max_row + 1):
                if col == 1:  # Column 1
                    sheet.cell(row=row, column=col).fill = fills[0]
                elif col == 2:  # Column 2
                    sheet.cell(row=row, column=col).fill = fills[1]
                elif col in (5, 6):  # Columns 5 and 6
                    sheet.cell(row=row, column=col).fill = fills[2] if col == 5 else fills[3]
                elif col in (7, 8):  # Columns 7 and 8
                    sheet.cell(row=row, column=col).fill = fills[4] if col == 7 else fills[5]

        # Color the last two columns based on specified colors
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=9).fill = PatternFill(start_color="FF5f867a", end_color="FF5f867a", fill_type="solid")  # Darker for 9th column
            sheet.cell(row=row, column=10).fill = PatternFill(start_color="FF4b5d67", end_color="FF4b5d67", fill_type="solid")  # Darker for 10th column

        # Alternate row colors while preserving column colors
        # for row in range(2, sheet.max_row + 1):
        #     if row % 2 == 0:
        #         for col in range(1, 11):
        #             sheet.cell(row=row, column=col).fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

    # Save the modified workbook
    wb.save(file_path)

# Example usage
color_spreadsheet('./tt.xlsx')
